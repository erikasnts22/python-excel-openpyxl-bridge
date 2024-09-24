from openpyxl.workbook import Workbook
#from openpyxl import load_workbook

# Create a workbook object
wb = Workbook()

# Create an active worksheet
ws = wb.active

# Create worksheet title
ws.title = "Names and Colors"

# Create Python list of names
names = ["Dan", "April", "Neal", "Sara"]
colors = ["Blue", "Purple", "Green", "White"]
nums = [12, 39, 42, 21]

ws["A1"] = "Names"
ws["B1"] = "Colors"
ws["C1"] = "Favorite Number"

# Add names to worksheet
starting_row = 2

for name in names:
	ws.cell(row=starting_row, column=1).value = name
	starting_row += 1

# Add colors to worksheet
starting_row = 2

for color in colors:
	ws.cell(row=starting_row, column=2).value = color
	starting_row += 1

# Add numbers to worksheet
starting_row = 2

for number in nums:
	ws.cell(row=starting_row, column=3).value = number
	starting_row += 1

# Use a formula
ws['C6'] = "=SUM(C2:C5)"

# Save spreadsheet
wb.save('names.xlsx')
print("File was saved!")