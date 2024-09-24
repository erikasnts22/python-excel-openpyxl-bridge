from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
# wb = Workbook()

# load existing spreadsheet
wb = load_workbook('hello.xlsx')

# Create an active worksheet
ws = wb.active

# Set a variable
#name = ws['A2'].value
#color = ws['B2'].value


# print something from the spreadsheet
#print(f'{name}: {color}')

# Grab a whole column
#column_a = ws['5']
#print(column_a)

# For loop
#for cell in column_a:
#	print(cell.value)

# Grab a range
#range = ws['A2:A10']
#print(range)

#for cell in range:
#	for x in cell:
#		print(x.value)

# Iterate thru columns
#for col in ws.iter_cols(min_row=1, max_row=10, min_col=1, max_col=2, values_only=True):
#	for cell in col:
#		print(cell)

# Change a cell
# ws['A2'] = "Johnny"

# Create Python list of names
names = ["Dan", "April", "Neal"]
#color = 

# Change many cells
starting_row = 11

for name in names:
	ws.cell(row=starting_row, column=1).value = name
	starting_row += 1

#ws.cell(row=starting_row, column=2).value = "Black"

# Save an excel spreadsheet
wb.save('hello2.xlsx')
print("File was saved!")