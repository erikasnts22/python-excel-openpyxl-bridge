from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# Create a workbook object
wb = Workbook()

# Create an active worksheet
ws = wb.active

# Select cell
cell = ws['A1']
cell2 = ws['B1']
cell3 = ws['C1']

# Change Font

cell.font = Font(
	size=30,
	bold=True,
	italic=False,
	color="a8329d")

cell2.font = Font(
	size=30,
	bold=True,
	italic=False,
	color="3250a8")

cell3.font = Font(
	size=30,
	bold=True,
	italic=True,
	color="bfb528")

# Define a side for border
my_bd = Side(
	style="double",
	color="000000")

B3 = ws['B3']

# Set border
B3.border = Border(
	left=my_bd,
	right=my_bd,
	top=my_bd,
	bottom=my_bd)

cell.border = Border(bottom=my_bd)
cell2.border = Border(bottom=my_bd)
cell3.border = Border(bottom=my_bd)



# Save spreadsheet
wb.save('names2.xlsx')
print("File was saved!")